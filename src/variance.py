"""
variance.py — Component 3: Budget vs. Actuals variance engine (FY2025).

WHAT FP&A CARES ABOUT HERE:
    A budget is a promise made in December; actuals are what reality delivered.
    The variance analysis is the post-mortem a CFO reads every month: where did
    we beat the plan, where did we miss, and *why*. The deliverables are:

      1. A line-by-line variance table ($ and %) for revenue, COGS, gross
         profit, each OpEx line, total OpEx, and operating income.
      2. Favorable / Unfavorable flags that respect sign convention — beating
         plan on revenue is good, but beating plan on spend means OVERspending.
      3. Material-variance flags (|%| > threshold or |$| > threshold) so the
         eye goes straight to what moved the needle.
      4. A variance BRIDGE (waterfall): plan operating income -> actual
         operating income, decomposed into the drivers of the gap.

OUTPUTS:
    outputs/variance_analysis.xlsx        (Summary + Monthly Detail + Bridge)
    outputs/figures/variance_bridge.png   (operating income waterfall)
    outputs/figures/variance_by_line.png  (favorable/unfavorable by line)

Run standalone:  python src/variance.py
"""
from __future__ import annotations

import sys

import numpy as np
import pandas as pd

import utils

# Windows consoles default to cp1252 and choke on the bullets/symbols we print.
# Force UTF-8 so the summary renders identically on every platform.
if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

# --- Materiality thresholds (config) ---------------------------------------
# A variance is "material" if it clears EITHER gate: a percentage gate catches
# small lines that swung hard, a dollar gate catches big lines that moved only
# a few percent. Both matter to a CFO.
MATERIAL_PCT = 0.05          # 5%
MATERIAL_DOLLAR = 250_000    # full-year $ threshold

# Line items in P&L reading order, with their "good direction".
# higher_better: actual above budget is FAVORABLE (revenue, profit lines)
# lower_better : actual below budget is FAVORABLE (cost lines)
LINE_SPECS = [
    ("revenue", "Revenue", "higher_better"),
    ("cogs", "COGS", "lower_better"),
    ("gross_profit", "Gross Profit", "higher_better"),
    ("sales_marketing", "Sales & Marketing", "lower_better"),
    ("research_development", "Research & Development", "lower_better"),
    ("general_admin", "General & Admin", "lower_better"),
    ("total_opex", "Total OpEx", "lower_better"),
    ("operating_income", "Operating Income", "higher_better"),
]


# ---------------------------------------------------------------------------
# Build the variance tables
# ---------------------------------------------------------------------------
def build_variance(actuals: pd.DataFrame, budget: pd.DataFrame):
    """
    Join FY2025 actuals to the FY2025 budget on month and compute variances at
    both the full-year and monthly grain. Returns (summary_df, monthly_df).
    """
    # The actuals file holds 24 months; the budget covers FY2025 only. The
    # inner join naturally isolates the 12 FY2025 months — no manual slicing.
    a = actuals.copy()
    a["gross_profit"] = a["revenue"] - a["cogs"]  # ensure present for the join

    b = budget.copy()
    b = b.rename(columns={c: c.replace("budget_", "") for c in b.columns if c != "month"})
    b["gross_profit"] = b["revenue"] - b["cogs"]  # budget gross profit (derived)

    line_cols = [c for c, _, _ in LINE_SPECS]
    merged = a.merge(b, on="month", suffixes=("_act", "_bud"), how="inner")
    merged = merged.sort_values("month").reset_index(drop=True)

    # --- full-year summary --------------------------------------------------
    summary_rows = []
    for col, label, direction in LINE_SPECS:
        bud = merged[f"{col}_bud"].sum()
        act = merged[f"{col}_act"].sum()
        var = act - bud
        var_pct = var / bud if bud else np.nan
        favorable = (var > 0) if direction == "higher_better" else (var < 0)
        material = abs(var_pct) >= MATERIAL_PCT or abs(var) >= MATERIAL_DOLLAR
        summary_rows.append({
            "Line Item": label,
            "Budget": bud,
            "Actual": act,
            "Variance $": var,
            "Variance %": var_pct,
            "F/U": "Favorable" if favorable else "Unfavorable",
            "Material": "Yes" if material else "",
        })
    summary = pd.DataFrame(summary_rows)

    # --- monthly detail (revenue, total opex, operating income drivers) -----
    monthly_rows = []
    for _, r in merged.iterrows():
        row = {"month": r["month"]}
        for col, label, direction in LINE_SPECS:
            var = r[f"{col}_act"] - r[f"{col}_bud"]
            row[f"{label} — Budget"] = r[f"{col}_bud"]
            row[f"{label} — Actual"] = r[f"{col}_act"]
            row[f"{label} — Var $"] = var
            row[f"{label} — Var %"] = var / r[f"{col}_bud"] if r[f"{col}_bud"] else np.nan
        monthly_rows.append(row)
    monthly = pd.DataFrame(monthly_rows)

    return summary, monthly, merged


def build_bridge(merged: pd.DataFrame) -> pd.DataFrame:
    """
    Decompose the operating-income gap (plan -> actual) into its drivers.

    Operating income = Revenue − COGS − S&M − R&D − G&A, so the total OI
    variance decomposes cleanly into one contribution per line. Revenue adds to
    OI when it beats plan; every cost line *subtracts* from OI when it runs over
    plan. Each bar's sign is its true contribution to the OI gap.
    """
    def s(col):  # full-year totals
        return merged[f"{col}_act"].sum() - merged[f"{col}_bud"].sum()

    budget_oi = merged["operating_income_bud"].sum()
    actual_oi = merged["operating_income_act"].sum()

    contributions = [
        ("Budget Operating Income", budget_oi, "anchor"),
        ("Revenue", s("revenue"), "delta"),                      # + helps OI
        ("COGS", -s("cogs"), "delta"),                           # overspend hurts
        ("Sales & Marketing", -s("sales_marketing"), "delta"),
        ("Research & Development", -s("research_development"), "delta"),
        ("General & Admin", -s("general_admin"), "delta"),
        ("Actual Operating Income", actual_oi, "anchor"),
    ]
    bridge = pd.DataFrame(contributions, columns=["Item", "Amount", "Kind"])

    # Sanity check: the deltas must reconcile budget OI -> actual OI exactly.
    delta_sum = bridge.loc[bridge["Kind"] == "delta", "Amount"].sum()
    assert abs((budget_oi + delta_sum) - actual_oi) < 1.0, "Bridge does not reconcile"
    return bridge


# ---------------------------------------------------------------------------
# Excel output
# ---------------------------------------------------------------------------
def write_excel(summary, monthly, bridge, path):
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter

    with pd.ExcelWriter(path, engine="openpyxl") as xl:
        summary.to_excel(xl, sheet_name="Summary", index=False, startrow=1)
        monthly_out = monthly.copy()
        monthly_out["month"] = monthly_out["month"].dt.strftime("%Y-%m")
        monthly_out.to_excel(xl, sheet_name="Monthly Detail", index=False)
        bridge.to_excel(xl, sheet_name="Bridge", index=False)

        wb = xl.book
        navy = "1B3A5B"
        green = "3F8F6B"
        red = "C0432F"
        header_fill = PatternFill("solid", fgColor=navy)
        header_font = Font(color="FFFFFF", bold=True)
        title_font = Font(color=navy, bold=True, size=14)
        fav_font = Font(color=green, bold=True)
        unfav_font = Font(color=red, bold=True)
        thin = Side(style="thin", color="DDDDDD")
        border = Border(bottom=thin)

        # --- Summary sheet formatting ---
        ws = wb["Summary"]
        ws["A1"] = "NorthPeak Analytics — FY2025 Budget vs. Actuals"
        ws["A1"].font = title_font
        # header row is row 2
        for cell in ws[2]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")
        # number formats + F/U coloring (data rows start at 3)
        money_cols = ["B", "C", "D"]   # Budget, Actual, Variance $
        for row in range(3, 3 + len(summary)):
            for col in money_cols:
                ws[f"{col}{row}"].number_format = '#,##0'
            ws[f"E{row}".format(row)].number_format = '0.0%'  # Variance %
            fu = ws[f"F{row}"]
            fu.font = fav_font if fu.value == "Favorable" else unfav_font
            for c in range(1, ws.max_column + 1):
                ws.cell(row=row, column=c).border = border
        # widen columns
        widths = {"A": 24, "B": 14, "C": 14, "D": 14, "E": 12, "F": 13, "G": 10}
        for col, w in widths.items():
            ws.column_dimensions[col].width = w
        ws.freeze_panes = "A3"

        # --- Monthly Detail: widen + number formats ---
        wsm = wb["Monthly Detail"]
        for cell in wsm[1]:
            cell.fill = header_fill
            cell.font = header_font
        wsm.column_dimensions["A"].width = 10
        for col_idx in range(2, wsm.max_column + 1):
            letter = get_column_letter(col_idx)
            wsm.column_dimensions[letter].width = 16
            header = wsm.cell(row=1, column=col_idx).value or ""
            fmt = '0.0%' if "Var %" in str(header) else '#,##0'
            for row in range(2, wsm.max_row + 1):
                wsm.cell(row=row, column=col_idx).number_format = fmt
        wsm.freeze_panes = "B2"

        # --- Bridge sheet ---
        wsb = wb["Bridge"]
        for cell in wsb[1]:
            cell.fill = header_fill
            cell.font = header_font
        wsb.column_dimensions["A"].width = 26
        wsb.column_dimensions["B"].width = 16
        wsb.column_dimensions["C"].width = 10
        for row in range(2, 2 + len(bridge)):
            wsb[f"B{row}"].number_format = '#,##0'


# ---------------------------------------------------------------------------
# Charts
# ---------------------------------------------------------------------------
def plot_bridge(bridge: pd.DataFrame, path):
    """Operating-income waterfall: plan -> actual, one bar per driver."""
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    utils.apply_style()
    fig, ax = plt.subplots(figsize=(12, 6.5))

    labels = bridge["Item"].tolist()
    amounts = bridge["Amount"].tolist()
    kinds = bridge["Kind"].tolist()

    running = 0.0
    for i, (label, amt, kind) in enumerate(zip(labels, amounts, kinds)):
        if kind == "anchor":
            ax.bar(i, amt, color=utils.COLORS["actual"], width=0.6, zorder=3)
            ax.text(i, amt + (0.01 * max(amounts)), utils.fmt_money(amt),
                    ha="center", va="bottom", fontsize=9, fontweight="bold")
            running = amt
        else:
            color = utils.COLORS["favorable"] if amt >= 0 else utils.COLORS["unfavorable"]
            bottom = running if amt >= 0 else running + amt
            ax.bar(i, abs(amt), bottom=bottom, color=color, width=0.6, zorder=3)
            # connector line from previous running total
            ax.plot([i - 1 + 0.3, i - 0.3], [running, running],
                    color="#999", lw=0.8, ls="--", zorder=2)
            sign = "+" if amt >= 0 else "−"
            ax.text(i, max(running, running + amt) + (0.01 * max(amounts)),
                    f"{sign}{utils.fmt_money(abs(amt))}", ha="center", va="bottom",
                    fontsize=8.5, color=color, fontweight="bold")
            running += amt

    ax.axhline(0, color="#444", lw=0.8)
    ax.set_xticks(range(len(labels)))
    ax.set_xticklabels(labels, rotation=20, ha="right", fontsize=9.5)
    ax.yaxis.set_major_formatter(FuncFormatter(utils.fmt_money))
    ax.set_title("FY2025 Operating Income Bridge — Budget vs. Actual")
    ax.set_ylabel("Operating Income")
    # custom legend
    from matplotlib.patches import Patch
    ax.legend(handles=[
        Patch(color=utils.COLORS["actual"], label="Plan / Actual (anchors)"),
        Patch(color=utils.COLORS["favorable"], label="Favorable to OI"),
        Patch(color=utils.COLORS["unfavorable"], label="Unfavorable to OI"),
    ], loc="upper right", fontsize=9)
    fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)


def plot_variance_by_line(summary: pd.DataFrame, path):
    """Horizontal bar of $ variance per line, colored favorable/unfavorable."""
    import matplotlib.pyplot as plt
    from matplotlib.ticker import FuncFormatter

    utils.apply_style()
    # Drop subtotal lines that double-count (gross profit, total opex) for clarity.
    show = summary[~summary["Line Item"].isin(["Gross Profit", "Total OpEx"])].copy()
    show = show.iloc[::-1]  # revenue on top
    colors = [utils.COLORS["favorable"] if fu == "Favorable"
              else utils.COLORS["unfavorable"] for fu in show["F/U"]]

    fig, ax = plt.subplots(figsize=(11, 6))
    ax.barh(show["Line Item"], show["Variance $"], color=colors, zorder=3)
    ax.axvline(0, color="#444", lw=0.8)
    ax.xaxis.set_major_formatter(FuncFormatter(utils.fmt_money))
    # Headroom on both sides so outboard data labels clear the bars and the
    # y-axis category labels (the negative Operating Income label needs room).
    vals = show["Variance $"]
    span = vals.max() - vals.min()
    ax.set_xlim(vals.min() - 0.28 * span, vals.max() + 0.18 * span)
    off = 0.012 * span
    for y, (v, mat) in enumerate(zip(show["Variance $"], show["Material"])):
        tag = "  ●" if mat == "Yes" else ""
        ax.text(v + np.sign(v) * off, y, f"{utils.fmt_money(v)}{tag}",
                va="center", ha="left" if v >= 0 else "right", fontsize=9)
    ax.set_title("FY2025 Variance by Line  (favorable = green, unfavorable = red, ● = material)")
    ax.set_xlabel("Actual − Budget ($)")
    fig.tight_layout()
    fig.savefig(path)
    plt.close(fig)


# ---------------------------------------------------------------------------
# Main
# ---------------------------------------------------------------------------
def main() -> None:
    utils.ensure_dirs()
    actuals = utils.load_actuals()
    budget = utils.load_budget()

    summary, monthly, merged = build_variance(actuals, budget)
    bridge = build_bridge(merged)

    xlsx_path = utils.OUTPUTS / "variance_analysis.xlsx"
    write_excel(summary, monthly, bridge, xlsx_path)
    plot_bridge(bridge, utils.FIGURES / "variance_bridge.png")
    plot_variance_by_line(summary, utils.FIGURES / "variance_by_line.png")

    # Console summary
    def row(label):
        return summary.loc[summary["Line Item"] == label].iloc[0]

    rev, oi = row("Revenue"), row("Operating Income")
    print("=" * 70)
    print("COMPONENT 3 — FY2025 BUDGET vs. ACTUALS VARIANCE")
    print("=" * 70)
    print(f"Period: {merged['month'].min():%Y-%m} to {merged['month'].max():%Y-%m} "
          f"({len(merged)} months)\n")
    print(f"{'Line Item':<24}{'Budget':>14}{'Actual':>14}{'Var $':>14}{'Var %':>9}  F/U")
    print("-" * 79)
    for _, r in summary.iterrows():
        mat = " ●" if r["Material"] == "Yes" else "  "
        print(f"{r['Line Item']:<24}{r['Budget']:>14,.0f}{r['Actual']:>14,.0f}"
              f"{r['Variance $']:>14,.0f}{r['Variance %']:>8.1%}  "
              f"{r['F/U'][:4]}{mat}")
    print("-" * 79)
    print(f"\nHeadline: Revenue {r_fu(rev)} {rev['Variance %']:+.1%} "
          f"({utils.fmt_money(rev['Variance $'])}); "
          f"Operating Income {r_fu(oi)} {utils.fmt_money(oi['Variance $'])} "
          f"vs plan.")
    print(f"\nWrote: outputs/variance_analysis.xlsx")
    print(f"Wrote: outputs/figures/variance_bridge.png")
    print(f"Wrote: outputs/figures/variance_by_line.png")
    print("=" * 70)


def r_fu(row) -> str:
    return "favorable" if row["F/U"] == "Favorable" else "unfavorable"


if __name__ == "__main__":
    main()
